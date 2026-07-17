"""Excel (.xlsx) parser using openpyxl."""

from __future__ import annotations

import re
import structlog
from pathlib import Path

from app.parser.base import BaseParser, ParsedRow, ParseResult

logger = structlog.get_logger()


class ExcelParser(BaseParser):
    """Parse BOQ tables from Excel (.xlsx) files using openpyxl.

    Strategy:
    1. Read all sheets (or the first sheet if only one)
    2. Detect header row by matching column names
    3. Map columns and extract data rows
    4. Handle merged cells and multi-line descriptions
    """

    def supports(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in (".xlsx", ".xls")

    def parse(self, file_path: Path) -> ParseResult:
        """Parse an Excel BOQ document."""
        from openpyxl import load_workbook

        result = ParseResult(source_file=str(file_path))

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            rows: list[ParsedRow] = []
            order_idx = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                col_mapping: dict[str, int] | None = None

                for row_data in ws.iter_rows(values_only=True):
                    # Convert to list, handle None values
                    cells = list(row_data)

                    if all(cell is None for cell in cells):
                        continue

                    # Detect header
                    if col_mapping is None:
                        mapping = self._detect_header(cells)
                        if mapping:
                            col_mapping = mapping
                            continue

                    if col_mapping:
                        parsed = self._extract_row(cells, col_mapping, order_idx)
                        if parsed and parsed.description.strip():
                            rows.append(parsed)
                            order_idx += 1

            wb.close()
            result.rows = rows
            result.total_pages = len(wb.sheetnames)
            logger.info(
                "excel_parse_complete",
                file=file_path.name,
                rows=len(rows),
                sheets=result.total_pages,
            )
        except Exception as e:
            result.error = f"Excel parsing failed: {e!s}"
            logger.error("excel_parse_error", file=str(file_path), error=str(e))

        return result

    def _detect_header(self, row: list) -> dict[str, int] | None:
        """Detect column header row."""
        cells = [str(cell).strip().lower() if cell else "" for cell in row]
        mapping: dict[str, int] = {}
        matches = 0

        for idx, cell in enumerate(cells):
            if not cell:
                continue
            if re.search(r"s\.?\s*no|sr\.?\s*no|sl\.?\s*no", cell):
                mapping["serial"] = idx
                matches += 1
            elif re.search(r"item\s*no|item\s*number|dsr\s*no", cell):
                mapping["item_number"] = idx
                matches += 1
            elif re.search(r"description|particulars|item\s*of\s*work", cell):
                mapping["description"] = idx
                matches += 1
            elif re.search(r"quant|qty", cell):
                mapping["quantity"] = idx
                matches += 1
            elif re.search(r"^unit$", cell):
                mapping["unit"] = idx
                matches += 1
            elif re.search(r"^rate$|rate\s*per", cell):
                mapping["rate"] = idx
                matches += 1
            elif re.search(r"amount|total|value", cell):
                mapping["amount"] = idx
                matches += 1

        if "description" in mapping and matches >= 3:
            return mapping
        return None

    def _extract_row(
        self, row: list, mapping: dict[str, int], order_idx: int
    ) -> ParsedRow | None:
        """Extract a ParsedRow from an Excel row."""
        def get_cell(key: str) -> str | None:
            idx = mapping.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else None
            return None

        description = get_cell("description") or ""
        if not description or len(description) < 2:
            return None

        item_num = get_cell("item_number") or get_cell("serial")
        depth = 0
        if item_num:
            parts = item_num.split(".")
            if len(parts) > 1:
                depth = len(parts) - 1

        return ParsedRow(
            order_index=order_idx,
            item_number=item_num,
            description=description,
            quantity=ParsedRow.parse_decimal(get_cell("quantity")),
            unit=get_cell("unit"),
            rate=ParsedRow.parse_decimal(get_cell("rate")),
            amount=ParsedRow.parse_decimal(get_cell("amount")),
            depth=depth,
        )
